import csv
import io
import json
import os
import socket
from datetime import datetime, timezone

from flask import Flask, jsonify, request, send_file, send_from_directory

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.environ.get("DATA_DIR", os.path.join(BASE_DIR, "data"))
CSV_PATH = os.path.join(DATA_DIR, "submissions.csv")
PDF_DIR = os.path.join(DATA_DIR, "pdfs")

FIELD_COLUMNS = [
    "submitted_at",
    "full_name",
    "date_of_birth",
    "gender",
    "occupation",
    "religion",
    "caste_community",
    "additional_identity",
    "mobile",
    "whatsapp",
    "state",
    "district",
    "parliament",
    "assembly",
    "village",
    "panchayat",
    "party_affiliation",
    "krishna_sir_follower",
    "position_role",
    "area_of_influence",
    "last_interaction_date",
    "grievances_1",
    "grievances_2",
    "grievances_3",
    "pdf_file",
]

REQUIRED_FIELDS = [
    "full_name",
    "date_of_birth",
    "gender",
    "occupation",
    "religion",
    "caste_community",
    "mobile",
    "whatsapp",
    "village",
    "party_affiliation",
    "krishna_sir_follower",
    "position_role",
    "area_of_influence",
]

EXPORT_KEY = os.environ.get("EXPORT_KEY", "")

app = Flask(__name__, static_folder=BASE_DIR, static_url_path="")


def ensure_csv():
    os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.isfile(CSV_PATH):
        with open(CSV_PATH, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=FIELD_COLUMNS)
            writer.writeheader()
        return

    with open(CSV_PATH, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        if reader.fieldnames and "pdf_file" in reader.fieldnames:
            return

    with open(CSV_PATH, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FIELD_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({col: row.get(col, "") for col in FIELD_COLUMNS})


def validate_submission(data):
    errors = []
    for field in REQUIRED_FIELDS:
        value = (data.get(field) or "").strip()
        if not value:
            errors.append(f"{field.replace('_', ' ').title()} is required.")

    mobile = (data.get("mobile") or "").strip()
    whatsapp = (data.get("whatsapp") or "").strip()
    if mobile and (not mobile.isdigit() or len(mobile) != 10):
        errors.append("Mobile Number must be exactly 10 digits.")
    if whatsapp and (not whatsapp.isdigit() or len(whatsapp) != 10):
        errors.append("WhatsApp Number must be exactly 10 digits.")

    return errors


def append_submission(data, pdf_filename=""):
    ensure_csv()
    row = {col: "" for col in FIELD_COLUMNS}
    row["submitted_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    for col in FIELD_COLUMNS:
        if col == "submitted_at":
            continue
        row[col] = (data.get(col) or "").strip()
    row["pdf_file"] = pdf_filename

    with open(CSV_PATH, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FIELD_COLUMNS)
        writer.writerow(row)


def save_pdf_upload(upload, full_name):
    os.makedirs(PDF_DIR, exist_ok=True)
    safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in (full_name or "registration"))[:40]
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"{safe_name}_{timestamp}.pdf"
    path = os.path.join(PDF_DIR, filename)
    upload.save(path)
    return filename, path


def normalize_header(name):
    return (name or "").strip().lower().replace(" ", "_").replace("/", "_")


def parse_csv_bytes(content):
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return None, "CSV file has no data rows."
    return rows, None


def parse_excel_bytes(content):
    if load_workbook is None:
        return None, "Excel support requires openpyxl. Run: pip install openpyxl"
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    sheet_rows = list(ws.iter_rows(values_only=True))
    if len(sheet_rows) < 2:
        return None, "Excel file has no data rows."
    headers = [normalize_header(str(h) if h is not None else "") for h in sheet_rows[0]]
    rows = []
    for values in sheet_rows[1:]:
        if not values or all(v is None or str(v).strip() == "" for v in values):
            continue
        row = dict(
            zip(
                headers,
                [str(v).strip() if v is not None else "" for v in values],
            )
        )
        rows.append(row)
    if not rows:
        return None, "Excel file has no data rows."
    return rows, None


def map_import_row(raw):
    mapped = {}
    alias_map = {
        "mobile_number_primary": "mobile",
        "mobile_number": "mobile",
        "whatsapp_number": "whatsapp",
        "caste_community": "caste_community",
        "party_affiliation": "party_affiliation",
        "krishna_sir_follower": "krishna_sir_follower",
        "position_role": "position_role",
        "area_of_influence": "area_of_influence",
        "last_interaction_date": "last_interaction_date",
        "open_grievances_1": "grievances_1",
        "open_grievances_2": "grievances_2",
        "open_grievances_3": "grievances_3",
        "open_grievances_requests_1": "grievances_1",
        "parliament_constituency": "parliament",
        "lok_sabha": "parliament",
        "assembly_constituency": "assembly",
    }
    for key, value in raw.items():
        norm = normalize_header(key)
        norm = alias_map.get(norm, norm)
        if norm in FIELD_COLUMNS and norm != "submitted_at":
            mapped[norm] = (value or "").strip()
    return mapped


def get_local_ip():
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        sock.connect(("8.8.8.8", 80))
        return sock.getsockname()[0]
    except OSError:
        return "127.0.0.1"
    finally:
        sock.close()


@app.after_request
def no_cache(response):
    if response.content_type and "text/html" in response.content_type:
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/data/<path:filename>")
def data_files(filename):
    return send_from_directory(os.path.join(BASE_DIR, "data"), filename)


@app.route("/css/<path:filename>")
def css_files(filename):
    return send_from_directory(os.path.join(BASE_DIR, "css"), filename)


@app.route("/js/<path:filename>")
def js_files(filename):
    return send_from_directory(os.path.join(BASE_DIR, "js"), filename)


@app.route("/api/health")
def health():
    return jsonify({"status": "ok"})


@app.route("/api/template")
def download_template():
    ensure_csv()
    output = io.StringIO()
    template_cols = [c for c in FIELD_COLUMNS if c != "submitted_at"]
    writer = csv.DictWriter(output, fieldnames=template_cols)
    writer.writeheader()
    writer.writerow({col: "" for col in template_cols})
    mem = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    mem.seek(0)
    return send_file(
        mem,
        mimetype="text/csv",
        as_attachment=True,
        download_name="registration_template.csv",
    )


def check_export_key():
    if not EXPORT_KEY:
        return False, jsonify({
            "success": False,
            "errors": ["Export not configured. Set EXPORT_KEY in Render Environment."],
        }), 503
    key = request.args.get("key", "")
    if key != EXPORT_KEY:
        return False, jsonify({"success": False, "errors": ["Invalid export key."]}), 403
    return True, None, None


@app.route("/api/export/submissions")
def export_submissions():
    ok, response, status = check_export_key()
    if not ok:
        return response, status

    ensure_csv()
    if not os.path.isfile(CSV_PATH):
        return jsonify({"success": False, "errors": ["No submissions yet."]}), 404

    with open(CSV_PATH, "r", encoding="utf-8-sig") as f:
        row_count = sum(1 for _ in f) - 1
    if row_count < 1:
        return jsonify({"success": False, "errors": ["No submissions yet."]}), 404

    return send_file(
        CSV_PATH,
        mimetype="text/csv",
        as_attachment=True,
        download_name="submissions.csv",
    )


@app.route("/api/export/count")
def export_count():
    ok, response, status = check_export_key()
    if not ok:
        return response, status

    ensure_csv()
    if not os.path.isfile(CSV_PATH):
        return jsonify({"count": 0})

    with open(CSV_PATH, "r", encoding="utf-8-sig") as f:
        rows = sum(1 for _ in f) - 1
    return jsonify({"count": max(rows, 0), "csv_path": CSV_PATH})


@app.route("/api/submit", methods=["POST"])
def submit_form():
    pdf_filename = ""
    pdf_path = ""

    if request.content_type and "multipart/form-data" in request.content_type:
        raw_data = request.form.get("data", "{}")
        try:
            data = json.loads(raw_data)
        except json.JSONDecodeError:
            return jsonify({"success": False, "errors": ["Invalid form data."]}), 400
        pdf_upload = request.files.get("pdf")
        if pdf_upload and pdf_upload.filename:
            try:
                pdf_filename, pdf_path = save_pdf_upload(pdf_upload, data.get("full_name", ""))
            except OSError as exc:
                return jsonify({"success": False, "errors": [f"Could not save PDF: {exc}"]}), 500
    else:
        data = request.get_json(silent=True) or {}

    for phone_field in ("mobile", "whatsapp"):
        if phone_field in data:
            data[phone_field] = "".join(ch for ch in str(data[phone_field]) if ch.isdigit())

    errors = validate_submission(data)
    if errors:
        return jsonify({"success": False, "errors": errors}), 400

    try:
        append_submission(data, pdf_filename)
    except OSError as exc:
        return jsonify({"success": False, "errors": [f"Could not save submission: {exc}"]}), 500

    return jsonify({
        "success": True,
        "message": "Registration submitted successfully. PDF and CSV saved.",
        "csv": CSV_PATH,
        "pdf": pdf_path or None,
        "pdf_file": pdf_filename or None,
    })


@app.route("/api/import", methods=["POST"])
def import_form():
    upload = request.files.get("file")
    if not upload or not upload.filename:
        return jsonify({"success": False, "errors": ["No file uploaded."]}), 400

    filename = upload.filename.lower()
    content = upload.read()

    if filename.endswith(".csv"):
        raw_rows, err = parse_csv_bytes(content)
    elif filename.endswith((".xlsx", ".xlsm")):
        raw_rows, err = parse_excel_bytes(content)
    elif filename.endswith(".xls"):
        return jsonify({
            "success": False,
            "errors": ["Legacy .xls is not supported. Please save as .xlsx or .csv."],
        }), 400
    else:
        return jsonify({
            "success": False,
            "errors": ["Unsupported file type. Upload .csv or .xlsx."],
        }), 400

    if err:
        return jsonify({"success": False, "errors": [err]}), 400

    mapped_rows = []
    row_warnings = []
    for index, raw in enumerate(raw_rows, start=2):
        row_dict = raw if isinstance(raw, dict) else {normalize_header(k): v for k, v in raw.items()}
        if not any(str(v).strip() for v in row_dict.values()):
            continue
        mapped = map_import_row(row_dict)
        if not mapped:
            row_warnings.append(f"Row {index}: no matching columns — skipped.")
            continue
        if not (mapped.get("full_name") or "").strip():
            row_warnings.append(f"Row {index}: missing full name — skipped.")
            continue
        if not (mapped.get("state") or "").strip():
            mapped["state"] = "Andhra Pradesh"
        mapped_rows.append(mapped)

    if not mapped_rows:
        return jsonify({
            "success": False,
            "errors": row_warnings or [
                "No valid candidate rows found. Download the template and use those column headers."
            ],
        }), 400

    return jsonify({
        "success": True,
        "rows": mapped_rows,
        "count": len(mapped_rows),
        "warnings": row_warnings,
    })


def find_free_port(preferred=8080):
    for port in [preferred, 5000, 8081, 8888, 9000]:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            try:
                sock.bind(("0.0.0.0", port))
                return port
            except OSError:
                continue
    raise RuntimeError("No free port found. Close other Python servers and retry.")


def run_server():
    ensure_csv()
    preferred = int(os.environ.get("PORT", 8080))
    port = find_free_port(preferred)
    local_ip = get_local_ip()
    print("=" * 56)
    print("Political Influencer Registration Form")
    print("=" * 56)
    if port != preferred:
        print(f"  NOTE: Port {preferred} was busy — using port {port} instead.")
    print(f"  Local:   http://127.0.0.1:{port}/")
    print(f"  Network: http://{local_ip}:{port}/")
    print(f"  CSV:     {CSV_PATH}")
    print(f"  PDFs:    {PDF_DIR}")
    print("=" * 56)
    print("Open the URL above in Chrome (desktop app), NOT in Cursor preview.")
    print("Do NOT open index.html directly (file:// will not work).")
    print("=" * 56)
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True, use_reloader=False)


if __name__ == "__main__":
    run_server()
